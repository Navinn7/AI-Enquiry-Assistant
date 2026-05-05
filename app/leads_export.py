"""
leads_export.py  –  Save leads to Excel file automatically.
"""
from pathlib import Path
from datetime import datetime

EXCEL_PATH = Path(__file__).parent.parent / "leads.xlsx"


def save_to_excel(telegram_id: str, name: str, interest: str, email: str, user_phone: str):
    try:
        import openpyxl

        if not EXCEL_PATH.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads"
            ws.append(["Timestamp", "Name", "Phone Number", "Email", "Course Interest", "Telegram ID", "Follow Up Sent"])
            wb.save(EXCEL_PATH)

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            name,
            user_phone,
            email,
            interest,
            telegram_id,
            "No",    # Follow Up Sent defaults to No
        ])
        wb.save(EXCEL_PATH)
        print(f"[EXCEL] Lead saved to {EXCEL_PATH}")

    except Exception as e:
        print(f"[EXCEL ERROR] {e}")


def mark_followup_sent(telegram_id: str):
    """Update Follow Up Sent column to Yes for this telegram_id."""
    try:
        import openpyxl

        if not EXCEL_PATH.exists():
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        # Find Telegram ID column index
        headers = [cell.value for cell in ws[1]]
        try:
            tid_col   = headers.index("Telegram ID") + 1
            fup_col   = headers.index("Follow Up Sent") + 1
        except ValueError:
            print("[EXCEL] Column not found")
            return

        # Find the row with this telegram_id and update
        for row in ws.iter_rows(min_row=2):
            if str(row[tid_col - 1].value) == str(telegram_id):
                row[fup_col - 1].value = "Yes"
                break

        wb.save(EXCEL_PATH)
        print(f"[EXCEL] Follow Up Sent marked for {telegram_id}")

    except Exception as e:
        print(f"[EXCEL ERROR] {e}")


def get_pending_followups() -> list[dict]:
    """Return all leads where Follow Up Sent is No."""
    try:
        import openpyxl

        if not EXCEL_PATH.exists():
            return []

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        try:
            name_col  = headers.index("Name") + 1
            tid_col   = headers.index("Telegram ID") + 1
            fup_col   = headers.index("Follow Up Sent") + 1
            ts_col    = headers.index("Timestamp") + 1
        except ValueError:
            return []

        pending = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[fup_col - 1] == "No":
                pending.append({
                    "name":        row[name_col - 1],
                    "telegram_id": str(row[tid_col - 1]),
                    "timestamp":   row[ts_col - 1],
                })
        return pending

    except Exception as e:
        print(f"[EXCEL ERROR] {e}")
        return []